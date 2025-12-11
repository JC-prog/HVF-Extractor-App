from openpyxl import Workbook

class XLSXExporter:
    def save(self, data: dict, out_path: str):
        wb = Workbook()
        ws = wb.active
        ws.title = "HVF Data"

        for row_idx, (key, value) in enumerate(data.items(), start=1):
            ws.cell(row=row_idx, column=1).value = key
            ws.cell(row=row_idx, column=2).value = value

        wb.save(out_path)
